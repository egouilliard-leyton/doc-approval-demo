"""Spreadsheet-mode (xlsx) generation tests: enumerate / read / list catalogue / fill.

Fully offline — pure openpyxl, no LibreOffice, no network. The LibreOffice recompute path
lives in ``test_generation_xlsx_preview.py`` (gated on ``soffice``).
"""

from __future__ import annotations

import shutil
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app import storage
from app.models import _new_id
from app.pipeline.generation import (
    enumerate_workbook_sheets,
    fill_spreadsheet,
    list_field_catalogue,
    read_computed_grid,
    read_template_grid,
    recompute_workbook,
)

from .generation_fixtures import make_xlsx_template

_HAS_SOFFICE = shutil.which("soffice") is not None


def _fv(value):
    return {"value": value, "confidence": 0.9, "grounding": None}


# A dumped InvoiceFields-shaped blob with two line items (drives the table row modes).
STRUCTURED_FIELDS = {
    "vendor": _fv("Acme Supplies Inc."),
    "total": _fv(135.0),
    "po_number": _fv(None),  # absent -> a bound scalar resolving here is skipped
    "line_items": [
        {"desc": _fv("Widget"), "qty": _fv(2), "unit_price": _fv(50.0), "amount": _fv(100.0)},
        {"desc": _fv("Gadget"), "qty": _fv(3), "unit_price": _fv(20.0), "amount": _fv(60.0)},
    ],
}


def _template_with_source(cell_map: dict) -> object:
    """A SimpleNamespace template ORM stand-in with an on-disk xlsx source."""
    tid = _new_id()
    storage.save_template_source(tid, ".xlsx", make_xlsx_template())
    return SimpleNamespace(id=tid, source_ext=".xlsx", cell_map=cell_map)


def _load(content: bytes):
    return load_workbook(BytesIO(content))["Invoice"]


# --- enumerate / read ---------------------------------------------------------


def test_enumerate_workbook_sheets(tmp_path):
    path = tmp_path / "tmpl.xlsx"
    path.write_bytes(make_xlsx_template())

    sheets = enumerate_workbook_sheets(path)
    assert [s.name for s in sheets] == ["Invoice"]
    meta = sheets[0]
    assert meta.max_row >= 6  # header row 3, anchor row 4, total row 6
    assert meta.max_col >= 4  # A..D


def test_read_template_grid_shows_formula_strings(tmp_path):
    path = tmp_path / "tmpl.xlsx"
    path.write_bytes(make_xlsx_template())

    grid = read_template_grid(path, "Invoice")
    by_addr = {c.address: c for c in grid.cells}
    # Labels read as display strings; empty cells are omitted.
    assert by_addr["A1"].value == "Vendor"
    assert "B1" not in by_addr
    # The per-row formula surfaces verbatim, flagged is_formula (no cached result).
    assert by_addr["D4"].is_formula is True
    assert by_addr["D4"].value == "=B4*C4"


def test_read_template_grid_missing_sheet_is_empty(tmp_path):
    path = tmp_path / "tmpl.xlsx"
    path.write_bytes(make_xlsx_template())
    grid = read_template_grid(path, "NoSuchSheet")
    assert grid.cells == []


# --- list catalogue -----------------------------------------------------------


def test_list_field_catalogue_composite_and_scalar():
    invoice = {e.list_path: e for e in list_field_catalogue("invoice")}
    assert "line_items" in invoice
    # A list_composite exposes its sub-model's record-relative columns.
    cols = {c.path for c in invoice["line_items"].columns}
    assert {"desc", "qty", "unit_price", "amount"} <= cols

    contract = {e.list_path: e for e in list_field_catalogue("contract")}
    assert "parties" in contract
    # A list_scalar exposes a single sentinel column (the record's own value).
    assert [c.path for c in contract["parties"].columns] == [""]


# --- scalar fill --------------------------------------------------------------


def test_fill_scalar_values_and_number_format_suffix():
    cell_map = {
        "scalars": [
            {"sheet": "Invoice", "cell": "B1", "field_path": "vendor", "suffix": None},
            {"sheet": "Invoice", "cell": "B2", "field_path": "total", "suffix": "USD"},
            {"sheet": "Invoice", "cell": "B10", "field_path": "po_number", "suffix": None},
        ],
        "tables": [],
    }
    tmpl = _template_with_source(cell_map)
    try:
        outcome = fill_spreadsheet(tmpl, STRUCTURED_FIELDS)
        ws = _load(outcome.content)

        # Text scalar: plain string, no format.
        assert ws["B1"].value == "Acme Supplies Inc."
        # Numeric scalar with a suffix: value stays NUMERIC, suffix rides in number_format.
        assert ws["B2"].value == 135.0
        assert ws["B2"].number_format == '#,##0.00" USD"'
        # A None value is skipped with a warning, never guessed (empty cell stays empty).
        assert ws["B10"].value is None
        assert "Invoice!B10" in outcome.skipped
        assert any("po_number" in w for w in outcome.warnings)
    finally:
        storage.delete_template_dir(tmpl.id)


# --- table fill: both row modes ----------------------------------------------


def test_fill_table_insert_row_clones_style_and_translates_formula():
    cell_map = {
        "scalars": [],
        "tables": [
            {
                "sheet": "Invoice",
                "list_path": "line_items",
                "anchor_cell": "A4",
                "row_mode": "insert_row",
                "columns": [
                    {"order": 0, "col": "A", "field_path": "desc"},
                    {"order": 1, "col": "B", "field_path": "qty"},
                    {"order": 2, "col": "C", "field_path": "unit_price", "suffix": "USD"},
                ],
            }
        ],
    }
    tmpl = _template_with_source(cell_map)
    try:
        outcome = fill_spreadsheet(tmpl, STRUCTURED_FIELDS)
        ws = _load(outcome.content)

        # Two records -> two rows (anchor row 4 filled in place + one inserted row 5).
        assert ws["A4"].value == "Widget"
        assert ws["A5"].value == "Gadget"
        # The inserted row cloned the anchor row's bold style.
        assert ws["A5"].font.bold is True
        # The unmapped per-row formula was cloned AND translated to the new row.
        assert ws["D4"].value == "=B4*C4"
        assert ws["D5"].value == "=B5*C5"
        # Numeric column with a suffix keeps a numeric value + a unit number_format.
        assert ws["C5"].value == 20.0
        assert ws["C5"].number_format == '#,##0.00" USD"'
        # insert_row pushes the total below the anchor down (row 6 -> row 7).
        assert ws["A7"].value == "Total"
        # BUG FIX: the bounded total that ended at the anchor row auto-expands to cover the
        # inserted row — =SUM(D4:D4) -> =SUM(D4:D5) — so it no longer silently omits rows.
        # (Pre-fix this stayed =SUM(D4:D4) and this assertion fails.)
        assert ws["D7"].value == "=SUM(D4:D5)"
        # The fully-covered total is not flagged as a residual risk.
        assert outcome.warnings == []
    finally:
        storage.delete_template_dir(tmpl.id)


_INSERT_ROW_CELL_MAP = {
    "scalars": [],
    "tables": [
        {
            "sheet": "Invoice",
            "list_path": "line_items",
            "anchor_cell": "A4",
            "row_mode": "insert_row",
            "columns": [
                {"order": 0, "col": "A", "field_path": "desc"},
                {"order": 1, "col": "B", "field_path": "qty"},
                {"order": 2, "col": "C", "field_path": "unit_price"},
            ],
        }
    ],
}


@pytest.mark.skipif(not _HAS_SOFFICE, reason="LibreOffice (soffice) not installed")
def test_fill_table_insert_row_total_recomputes_over_inserted_rows():
    """The auto-expanded total, once recomputed by LibreOffice, equals the true sum.

    Two line items -> one inserted row; per-row D=B*C gives D4=2*50=100, D5=3*20=60. The
    widened =SUM(D4:D5) must recompute to 160 (pre-fix =SUM(D4:D4) would compute only 100).
    """
    tmpl = _template_with_source(_INSERT_ROW_CELL_MAP)
    try:
        outcome = fill_spreadsheet(tmpl, STRUCTURED_FIELDS)
        # The formula string was widened to span the inserted row.
        assert _load(outcome.content)["D7"].value == "=SUM(D4:D5)"

        result = recompute_workbook(tmpl.id, outcome.content)
        assert result.computed is True
        by_addr = {c.address: c for s in read_computed_grid(result.xlsx_bytes) for c in s.cells}
        # 100 (row 4) + 60 (row 5) == 160, the sum of BOTH rows' amounts.
        assert by_addr["D7"].value == "160"
    finally:
        storage.delete_template_dir(tmpl.id)


def test_fill_table_fill_next_empty_row_overwrites_without_inserting():
    cell_map = {
        "scalars": [],
        "tables": [
            {
                "sheet": "Invoice",
                "list_path": "line_items",
                "anchor_cell": "A4",
                "row_mode": "fill_next_empty_row",
                "columns": [{"order": 0, "col": "A", "field_path": "desc"}],
            }
        ],
    }
    tmpl = _template_with_source(cell_map)
    try:
        outcome = fill_spreadsheet(tmpl, STRUCTURED_FIELDS)
        ws = _load(outcome.content)

        assert ws["A4"].value == "Widget"
        assert ws["A5"].value == "Gadget"
        # No row was inserted, so the total stays put at row 6.
        assert ws["A6"].value == "Total"
    finally:
        storage.delete_template_dir(tmpl.id)


def test_fill_table_row_count_from_resolve_path():
    """N rows written == len(resolve_path(fields, list_path)) — here exactly 2."""
    cell_map = {
        "scalars": [],
        "tables": [
            {
                "sheet": "Invoice",
                "list_path": "line_items",
                "anchor_cell": "A4",
                "row_mode": "fill_next_empty_row",
                "columns": [{"order": 0, "col": "A", "field_path": "desc"}],
            }
        ],
    }
    tmpl = _template_with_source(cell_map)
    try:
        outcome = fill_spreadsheet(tmpl, STRUCTURED_FIELDS)
        written = [addr for addr in outcome.filled if addr.startswith("Invoice!A")]
        assert written == ["Invoice!A4", "Invoice!A5"]  # 2 records -> 2 rows
    finally:
        storage.delete_template_dir(tmpl.id)


def test_fill_table_non_list_source_warns_and_writes_nothing():
    """A list_path resolving to a non-list value yields 0 rows + a warning, never raises."""
    cell_map = {
        "scalars": [],
        "tables": [
            {
                "sheet": "Invoice",
                "list_path": "vendor",  # a scalar, not a list
                "anchor_cell": "A4",
                "row_mode": "fill_next_empty_row",
                "columns": [{"order": 0, "col": "A", "field_path": ""}],
            }
        ],
    }
    tmpl = _template_with_source(cell_map)
    try:
        outcome = fill_spreadsheet(tmpl, STRUCTURED_FIELDS)
        assert outcome.filled == []
        assert any("not a list" in w for w in outcome.warnings)
        # The anchor cell is untouched (still blank).
        assert _load(outcome.content)["A4"].value is None
    finally:
        storage.delete_template_dir(tmpl.id)
