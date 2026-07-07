"""Smoke tests: the two headline template->generate journeys, end to end.

These mirror the exact happy paths validated in a real browser with PinchTab:
  1. form-fill: create -> upload a fillable PDF -> AcroForm auto-detected ->
     AI/heuristic mapping suggestion -> save mapping -> generate a filled PDF.
  2. rich-html: create -> author an HTML body with a data-field placeholder ->
     generate PDF + DOCX with the placeholder bound to a document's value.

Everything runs offline through the public API (TestClient) with the mock
mapping heuristic and a stashed mock structure result -- no API key, no network.
They are deliberately coarse-grained golden-path guards; the fine-grained
behavior lives in the other test_generation_* modules.
"""

import io

import docx
from fastapi.testclient import TestClient
from pypdf import PdfReader
from sqlmodel import Session

from app import storage
from app.db import engine
from app.main import app
from app.models import PipelineRun, _new_id

from .generation_fixtures import make_fillable_pdf, make_xlsx_template


def _fv(value):
    return {"value": value, "confidence": 0.9, "grounding": None}


# A dumped InvoiceFields-shaped blob (FieldValue leaves keyed value/confidence/grounding).
STRUCTURED_FIELDS = {
    "vendor": _fv("MOCK INVOICE"),
    "total": _fv(1234.56),
    "currency": _fv("USD"),
    "line_items": [{"desc": _fv("Widget"), "amount": _fv(125.0)}],
}


def _stash_structured_doc() -> str:
    """Insert a document's latest PipelineRun carrying a mock structure result."""
    doc_id = _new_id()
    with Session(engine) as session:
        session.add(
            PipelineRun(
                document_id=doc_id,
                status="structured",
                stage_results={"structure": {"fields": STRUCTURED_FIELDS}},
            )
        )
        session.commit()
    return doc_id


def test_smoke_form_fill_journey():
    """Create -> upload fillable PDF -> AcroForm detected -> heuristic map -> generate."""
    with TestClient(app) as client:
        # Create an invoice template (the Phase 0 wizard's "Start blank").
        tid = client.post("/templates", json={"name": "PO Form", "doc_type": "invoice"}).json()["id"]

        # Upload a fillable PDF source -> the backend flips the template to form_fill
        # and enumerates its AcroForm widgets.
        detail = client.post(
            f"/templates/{tid}/source",
            files={"file": ("fillable.pdf", make_fillable_pdf(), "application/pdf")},
        ).json()
        assert detail["mode"] == "form_fill"
        field_names = {f["name"] for f in detail["form_fields"]}
        assert {"vendor_name", "total_amount", "currency", "Signature"} <= field_names

        # The offline heuristic suggests bindings; obvious names should resolve.
        suggest = client.post(f"/templates/{tid}/suggest-mapping")
        assert suggest.status_code == 200, suggest.text
        suggestions = suggest.json()["suggestions"]
        assert suggestions["vendor_name"]["field_path"] == "vendor"
        assert suggestions["currency"]["field_path"] == "currency"
        assert suggestions["Signature"]["is_signature"] is True

        # Persist the accepted mapping (what the "Save mapping" button does).
        field_map = {
            name: {
                "field_path": s["field_path"],
                "is_signature": s["is_signature"],
            }
            for name, s in suggestions.items()
        }
        assert client.put(f"/templates/{tid}", json={"form_field_map": field_map}).status_code == 200

        # Generate from a processed document.
        doc_id = _stash_structured_doc()
        gen = client.post(
            f"/templates/{tid}/generate", params={"document_id": doc_id, "flatten": True}
        )
        assert gen.status_code == 201, gen.text
        body = gen.json()
        assert {"vendor_name", "total_amount", "currency"} <= set(body["filled_fields"])
        assert body["outputs"] and body["outputs"][0]["format"] == "pdf"

        # The output PDF exists and its AcroForm carries the bound values.
        out = storage.template_outputs_dir(tid) / f"{body['output_id']}.pdf"
        assert out.exists()
        fields = PdfReader(str(out)).get_fields()
        assert fields["vendor_name"]["/V"] == "MOCK INVOICE"
        assert fields["total_amount"]["/V"] == "1234.56"
        assert fields["currency"]["/V"] == "USD"


def test_smoke_rich_html_journey():
    """Create -> author HTML with a data-field placeholder -> generate PDF + DOCX."""
    with TestClient(app) as client:
        tid = client.post(
            "/templates", json={"name": "Invoice Letter", "doc_type": "invoice"}
        ).json()["id"]

        # Author the body (what the TipTap editor's Save persists) and pick both formats.
        html = '<p>Invoice for <span data-field="vendor" data-field-kind="text">Vendor</span></p>'
        put = client.put(
            f"/templates/{tid}",
            json={"html_body": html, "output_formats": ["pdf", "docx"]},
        )
        assert put.status_code == 200, put.text

        doc_id = _stash_structured_doc()
        gen = client.post(f"/templates/{tid}/generate", params={"document_id": doc_id})
        assert gen.status_code == 201, gen.text
        body = gen.json()
        assert "vendor" in body["filled_fields"]

        outputs = {o["format"]: o for o in body["outputs"]}
        assert set(outputs) == {"pdf", "docx"}

        # The PDF renders and contains the bound vendor value.
        pdf_path = storage.template_outputs_dir(tid) / f"{outputs['pdf']['output_id']}.pdf"
        assert pdf_path.exists()
        assert pdf_path.read_bytes()[:4] == b"%PDF"
        assert "MOCK INVOICE" in PdfReader(str(pdf_path)).pages[0].extract_text()

        # The DOCX renders and contains the bound vendor value.
        docx_path = storage.template_outputs_dir(tid) / f"{outputs['docx']['output_id']}.docx"
        assert docx_path.exists()
        assert docx_path.read_bytes()[:2] == b"PK"
        text = "\n".join(p.text for p in docx.Document(io.BytesIO(docx_path.read_bytes())).paragraphs)
        assert "MOCK INVOICE" in text


def test_smoke_generate_and_sign_journey():
    """Create rich-html template -> generate PDF -> seal it with a PAdES signature.

    The outbound path validated in a real browser with PinchTab: a generated output
    PDF is signed via the mock provider (offline), the response self-validates, and a
    ``<output_id>-signed.pdf`` lands on disk beside the original output.
    """
    with TestClient(app) as client:
        tid = client.post(
            "/templates", json={"name": "Signable Letter", "doc_type": "invoice"}
        ).json()["id"]
        html = '<p>Invoice for <span data-field="vendor" data-field-kind="text">Vendor</span></p>'
        assert client.put(
            f"/templates/{tid}", json={"html_body": html, "output_formats": ["pdf"]}
        ).status_code == 200

        doc_id = _stash_structured_doc()
        gen = client.post(f"/templates/{tid}/generate", params={"document_id": doc_id})
        assert gen.status_code == 201, gen.text
        oid = gen.json()["output_id"]

        signed = client.post(
            f"/templates/{tid}/outputs/{oid}/sign", params={"provider": "mock"}
        )
        assert signed.status_code == 201, signed.text
        assert signed.json()["validation"]["valid"] is True

        signed_path = storage.template_outputs_dir(tid) / f"{oid}-signed.pdf"
        assert signed_path.exists()
        assert signed_path.read_bytes()[:4] == b"%PDF"


def test_smoke_spreadsheet_journey():
    """Create -> upload .xlsx -> spreadsheet mode -> map a scalar + a line-items table
    -> generate -> the mapped scalar value AND a table row value land in the right cells."""
    from openpyxl import load_workbook

    with TestClient(app) as client:
        tid = client.post(
            "/templates", json={"name": "Invoice Sheet", "doc_type": "invoice"}
        ).json()["id"]

        # Upload the .xlsx source -> the backend flips the template to spreadsheet mode
        # and enumerates its sheet layout.
        detail = client.post(
            f"/templates/{tid}/source",
            files={
                "file": ("invoice.xlsx", make_xlsx_template(), storage.XLSX_MIME),
            },
        ).json()
        assert detail["mode"] == "spreadsheet"
        assert [s["name"] for s in detail["spreadsheet_sheets"]] == ["Invoice"]

        # Persist a cell map: a scalar (vendor -> B1) and a line-items table anchored at
        # A4 (desc -> col A, amount -> col D), mirroring the click-to-bind UI's PUT.
        cell_map = {
            "scalars": [
                {"sheet": "Invoice", "cell": "B1", "field_path": "vendor",
                 "suffix": None, "is_signature": False},
            ],
            "tables": [
                {
                    "sheet": "Invoice",
                    "list_path": "line_items",
                    "anchor_cell": "A4",
                    "row_mode": "fill_next_empty_row",
                    "columns": [
                        {"order": 0, "col": "A", "field_path": "desc", "suffix": None},
                        {"order": 1, "col": "D", "field_path": "amount", "suffix": None},
                    ],
                }
            ],
        }
        put = client.put(f"/templates/{tid}", json={"cell_map": cell_map})
        assert put.status_code == 200, put.text

        # Generate the filled .xlsx from a processed document.
        doc_id = _stash_structured_doc()
        gen = client.post(f"/templates/{tid}/generate", params={"document_id": doc_id})
        assert gen.status_code == 201, gen.text
        body = gen.json()
        assert body["outputs"] and body["outputs"][0]["format"] == "xlsx"

        # Open the produced workbook and assert the scalar + a table row value landed.
        out = storage.template_outputs_dir(tid) / f"{body['output_id']}.xlsx"
        assert out.exists()
        wb = load_workbook(out, data_only=False)
        ws = wb["Invoice"]
        assert ws["B1"].value == "MOCK INVOICE"  # scalar binding
        assert ws["A4"].value == "Widget"  # first line-item's desc
        assert ws["D4"].value == 125.0  # first line-item's amount
