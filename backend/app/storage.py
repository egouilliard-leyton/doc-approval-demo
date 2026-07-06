"""Filesystem ingestion + normalization: save originals, rasterize to per-page PNGs.

On-disk layout per document::

    data/<doc_id>/
        original.<ext>
        pages/page-001.png      # full-res rasterized page
        thumbs/page-001.png     # downscaled preview

PDFs are rendered with PyMuPDF (fitz); images are normalized with Pillow.
"""

from __future__ import annotations

import csv
import json
import logging
import shutil
from pathlib import Path

import fitz  # PyMuPDF
import numpy as np
from PIL import Image, ImageSequence

from app.config import settings

logger = logging.getLogger(__name__)

# Spreadsheet MIME types. These take a separate ingest path: instead of rasterizing
# to page images they are parsed cell-by-cell (see ``_normalize_spreadsheet``) and the
# grid is grounded/rendered natively — a spreadsheet has no page image.
CSV_MIME = "text/csv"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SPREADSHEET_MIMES: frozenset[str] = frozenset({CSV_MIME, XLSX_MIME})

# Allowed upload types: extension -> canonical MIME.
ALLOWED_TYPES: dict[str, str] = {
    ".pdf": "application/pdf",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".tif": "image/tiff",
    ".tiff": "image/tiff",
    ".csv": CSV_MIME,
    ".xlsx": XLSX_MIME,
}

# Demo caps so a pathological workbook can't produce a giant grid. Truncation is
# recorded per sheet (``truncated_rows``/``truncated_cols``) and surfaced as an OCR
# warning by ``SpreadsheetEngine`` — never silently dropped.
MAX_SHEET_ROWS = 500
MAX_SHEET_COLS = 60


def is_spreadsheet(mime: str) -> bool:
    """True for upload MIMEs that take the native spreadsheet path (no rasterization)."""
    return mime in SPREADSHEET_MIMES


# Accepted template source uploads: extension -> canonical MIME. A template source is
# either a fillable PDF (form-fill mode) or a Word doc converted to rich HTML (Phase 2).
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
TEMPLATE_SOURCE_TYPES: dict[str, str] = {
    ".pdf": "application/pdf",
    ".docx": DOCX_MIME,
}


class UnsupportedFileType(Exception):
    """Raised when an upload has an extension we don't accept."""


def _doc_dir(doc_id: str) -> Path:
    return settings.data_path / doc_id


def detect_type(filename: str) -> tuple[str, str]:
    """Return (lowercased extension, MIME) for an accepted file, else raise."""
    ext = Path(filename).suffix.lower()
    mime = ALLOWED_TYPES.get(ext)
    if mime is None:
        raise UnsupportedFileType(ext or filename)
    return ext, mime


def detect_template_source_type(filename: str) -> tuple[str, str]:
    """Return (lowercased extension, MIME) for an accepted template source, else raise."""
    ext = Path(filename).suffix.lower()
    mime = TEMPLATE_SOURCE_TYPES.get(ext)
    if mime is None:
        raise UnsupportedFileType(ext or filename)
    return ext, mime


def delete_document_dir(doc_id: str) -> None:
    """Recursively remove data/<doc_id>/ (originals, pages, thumbs, stage artifacts).

    A missing directory (e.g. a doc whose normalization failed before any file was
    written) is treated as success, so deletion stays idempotent. Real failures
    (permissions, a locked file) are logged rather than silently swallowed — the DB
    row is already gone, so a leftover tree is an orphan worth surfacing.
    """
    doc_dir = _doc_dir(doc_id)
    if not doc_dir.exists():
        return
    shutil.rmtree(
        doc_dir,
        onexc=lambda _func, path, exc: logger.warning(
            "Failed to delete %s while removing document %s: %s", path, doc_id, exc
        ),
    )


def save_original(doc_id: str, ext: str, content: bytes) -> Path:
    """Persist the raw upload bytes as data/<doc_id>/original.<ext>."""
    doc_dir = _doc_dir(doc_id)
    doc_dir.mkdir(parents=True, exist_ok=True)
    original = doc_dir / f"original{ext}"
    original.write_bytes(content)
    return original


def _save_page(doc_id: str, page_no: int, image: Image.Image) -> None:
    """Write one normalized page PNG + its thumbnail."""
    doc_dir = _doc_dir(doc_id)
    pages_dir = doc_dir / "pages"
    thumbs_dir = doc_dir / "thumbs"
    pages_dir.mkdir(parents=True, exist_ok=True)
    thumbs_dir.mkdir(parents=True, exist_ok=True)

    rgb = image.convert("RGB")
    name = f"page-{page_no:03d}.png"
    rgb.save(pages_dir / name, "PNG")

    # The full-res page is already written; downscale ``rgb`` in place for the thumb
    # (height is left effectively unbounded so the aspect ratio is preserved).
    rgb.thumbnail((settings.thumbnail_width, settings.thumbnail_width * 10))
    rgb.save(thumbs_dir / name, "PNG")


def normalize_to_pages(doc_id: str, original: Path, mime: str) -> int:
    """Render the original into per-page PNGs (+ thumbnails). Returns page count.

    Spreadsheets take a separate path: they are parsed into ``sheets.json`` (one page
    per sheet) rather than rasterized, since a spreadsheet has no page image.
    """
    if mime == "application/pdf":
        return _normalize_pdf(doc_id, original)
    if is_spreadsheet(mime):
        return _normalize_spreadsheet(doc_id, original, mime)
    return _normalize_image(doc_id, original)


def _normalize_pdf(doc_id: str, original: Path) -> int:
    zoom = settings.render_dpi / 72.0  # PDF user space is 72 DPI.
    matrix = fitz.Matrix(zoom, zoom)
    with fitz.open(original) as pdf:
        for index, page in enumerate(pdf, start=1):
            pix = page.get_pixmap(matrix=matrix)
            image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            _save_page(doc_id, index, image)
        return pdf.page_count


def _normalize_image(doc_id: str, original: Path) -> int:
    """Single images become one page; multi-frame TIFFs become one page per frame."""
    page_no = 0
    with Image.open(original) as image:
        for frame in ImageSequence.Iterator(image):
            page_no += 1
            _save_page(doc_id, page_no, frame)
    return page_no


# --- spreadsheet ingestion (CSV/XLSX) ----------------------------------------


def _cell_to_str(value: object) -> str:
    """Render a parsed cell as a display string (empty for blanks).

    openpyxl with ``data_only=True`` hands back computed values; whole-number floats
    (``5.0``) are collapsed to ``"5"`` so amounts read like the sheet, not like Python.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _parse_csv(original: Path) -> list[dict]:
    """Parse a CSV into a single-sheet list, applying the row/col caps."""
    rows: list[list[str]] = []
    truncated_cols = False
    truncated_rows = False
    with original.open("r", encoding="utf-8", errors="replace", newline="") as fh:
        for i, raw in enumerate(csv.reader(fh)):
            if i >= MAX_SHEET_ROWS:
                truncated_rows = True
                break
            if len(raw) > MAX_SHEET_COLS:
                truncated_cols = True
                raw = raw[:MAX_SHEET_COLS]
            rows.append([_cell_to_str(c) for c in raw])
    return [
        {
            "name": "Sheet1",
            "rows": rows,
            "truncated_rows": truncated_rows,
            "truncated_cols": truncated_cols,
        }
    ]


def _parse_xlsx(original: Path) -> list[dict]:
    """Parse an XLSX workbook into per-sheet row grids, applying the row/col caps.

    ``read_only`` streams rows without loading the whole workbook; ``data_only`` uses
    the cached computed value of any formula (not the formula string). Merged cells
    surface their value in the top-left cell only — good enough for the demo.
    """
    from openpyxl import load_workbook  # lazy: keep import cost off app boot

    workbook = load_workbook(original, read_only=True, data_only=True)
    sheets: list[dict] = []
    try:
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            truncated_cols = False
            for r, row in enumerate(worksheet.iter_rows(values_only=True)):
                if r >= MAX_SHEET_ROWS:
                    break
                cells = list(row)
                if len(cells) > MAX_SHEET_COLS:
                    truncated_cols = True
                    cells = cells[:MAX_SHEET_COLS]
                rows.append([_cell_to_str(c) for c in cells])
            # Drop fully-empty trailing rows (openpyxl often over-reports the extent).
            while rows and all(c == "" for c in rows[-1]):
                rows.pop()
            truncated_rows = bool(worksheet.max_row and worksheet.max_row > MAX_SHEET_ROWS)
            sheets.append(
                {
                    "name": worksheet.title,
                    "rows": rows,
                    "truncated_rows": truncated_rows,
                    "truncated_cols": truncated_cols,
                }
            )
    finally:
        workbook.close()
    return sheets


def _normalize_spreadsheet(doc_id: str, original: Path, mime: str) -> int:
    """Parse a spreadsheet into ``sheets.json`` (one page per sheet). Returns sheet count.

    Writes the parsed grid to ``data/<doc_id>/sheets.json`` (served via /files) so the
    frontend grid renders immediately, independent of the OCR stage. No PNGs are written.
    """
    sheets = _parse_csv(original) if mime == CSV_MIME else _parse_xlsx(original)
    if not sheets:  # a workbook with no worksheets: keep one empty page
        sheets = [{"name": "Sheet1", "rows": [], "truncated_rows": False, "truncated_cols": False}]

    doc_dir = _doc_dir(doc_id)
    doc_dir.mkdir(parents=True, exist_ok=True)
    sheets_json_path(doc_id).write_text(json.dumps(sheets), encoding="utf-8")
    return len(sheets)


def sheets_json_path(doc_id: str) -> Path:
    """Absolute path to the parsed spreadsheet grid (one entry per sheet)."""
    return _doc_dir(doc_id) / "sheets.json"


def sheets_url(doc_id: str) -> str:
    """Relative URL (served via /files) for the parsed spreadsheet grid."""
    return f"/files/{doc_id}/sheets.json"


def page_urls(doc_id: str, page_count: int) -> list[dict[str, object]]:
    """Relative URLs (served via the /files static mount) for each page + thumbnail."""
    return [
        {
            "page": n,
            "image_url": f"/files/{doc_id}/pages/page-{n:03d}.png",
            "thumbnail_url": f"/files/{doc_id}/thumbs/page-{n:03d}.png",
        }
        for n in range(1, page_count + 1)
    ]


def page_path(doc_id: str, page_no: int) -> Path:
    """Absolute path to one rasterized full-res page PNG."""
    return _doc_dir(doc_id) / "pages" / f"page-{page_no:03d}.png"


# --- Phase 2: pre-flight preprocessing artifacts -----------------------------


def prescan_dir(doc_id: str) -> Path:
    """Directory holding cleaned (deskewed/grayscale/threshold) page variants."""
    return _doc_dir(doc_id) / "prescan"


def save_prescan_page(doc_id: str, page_no: int, variant: str, image: np.ndarray) -> Path:
    """Write a cleaned page variant as prescan/page-NNN-<variant>.png.

    Accepts an OpenCV ndarray: 2-D arrays are treated as grayscale, 3-D arrays
    as BGR (OpenCV's default channel order) and converted to RGB before saving.
    """
    out_dir = prescan_dir(doc_id)
    out_dir.mkdir(parents=True, exist_ok=True)

    if image.ndim == 2:
        pil = Image.fromarray(image, mode="L")
    else:
        pil = Image.fromarray(image[:, :, ::-1])  # BGR -> RGB

    path = out_dir / f"page-{page_no:03d}-{variant}.png"
    pil.save(path, "PNG")
    return path


def prescan_url(doc_id: str, page_no: int, variant: str) -> str:
    """Relative URL (served via /files) for a cleaned page variant."""
    return f"/files/{doc_id}/prescan/page-{page_no:03d}-{variant}.png"


# --- Phase 1 (signatures): detected signature crops --------------------------


def signatures_dir(doc_id: str) -> Path:
    """Directory holding cropped signature images from the detection post-pass."""
    return _doc_dir(doc_id) / "signatures"


def save_signature_crop(doc_id: str, page_no: int, index: int, image: Image.Image) -> Path:
    """Write one detected signature crop as signatures/page-NNN-sig-II.png.

    Accepts a PIL image (the crop taken from the page PNG); saved as RGB PNG.
    """
    out_dir = signatures_dir(doc_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"page-{page_no:03d}-sig-{index:02d}.png"
    image.convert("RGB").save(path, "PNG")
    return path


def signature_crop_url(doc_id: str, page_no: int, index: int) -> str:
    """Relative URL (served via /files) for a saved signature crop."""
    return f"/files/{doc_id}/signatures/page-{page_no:03d}-sig-{index:02d}.png"


# --- Phase 3: OCR artifacts (per-engine markdown) ----------------------------


def ocr_dir(doc_id: str, engine: str) -> Path:
    """Directory holding an engine's OCR artifacts, e.g. per-page markdown."""
    return _doc_dir(doc_id) / "ocr" / engine


def save_ocr_markdown(doc_id: str, engine: str, page_no: int, markdown: str) -> Path:
    """Write a page's OCR markdown as ocr/<engine>/page-NNN.md."""
    out_dir = ocr_dir(doc_id, engine)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"page-{page_no:03d}.md"
    path.write_text(markdown, encoding="utf-8")
    return path


def ocr_markdown_url(doc_id: str, engine: str, page_no: int) -> str:
    """Relative URL (served via /files) for a page's saved OCR markdown."""
    return f"/files/{doc_id}/ocr/{engine}/page-{page_no:03d}.md"


# --- Phase 4: structuring artifacts (raw extractor output) -------------------


def structure_dir(doc_id: str) -> Path:
    """Directory holding the structuring stage's raw extractor output."""
    return _doc_dir(doc_id) / "structure"


def save_structure_artifact(doc_id: str, content: str, name: str = "extractions.jsonl") -> Path:
    """Persist the raw extractor output (e.g. LangExtract JSONL) for debugging/demo."""
    out_dir = structure_dir(doc_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / name
    path.write_text(content, encoding="utf-8")
    return path


def structure_artifact_url(doc_id: str, name: str = "extractions.jsonl") -> str:
    """Relative URL (served via /files) for the saved structuring artifact."""
    return f"/files/{doc_id}/structure/{name}"


# --- Phase 1 (form-fill): template source + generated outputs -----------------

# Templates get their own data/templates/<id>/ tree so their source PDFs and
# generated outputs never collide with a document's data/<doc_id>/.


def _template_dir(template_id: str) -> Path:
    return settings.data_path / "templates" / template_id


def template_dir(template_id: str) -> Path:
    """Directory holding a template's source PDF + generated outputs."""
    return _template_dir(template_id)


def delete_template_dir(template_id: str) -> None:
    """Recursively remove data/templates/<template_id>/ (source + outputs).

    A missing directory (e.g. a template that never got a source upload) is treated
    as success, so deletion stays idempotent. Real failures (permissions, a locked
    file) are logged rather than silently swallowed — the DB row is already gone, so
    a leftover tree is an orphan worth surfacing.
    """
    tmpl_dir = _template_dir(template_id)
    if not tmpl_dir.exists():
        return
    shutil.rmtree(
        tmpl_dir,
        onexc=lambda _func, path, exc: logger.warning(
            "Failed to delete %s while removing template %s: %s", path, template_id, exc
        ),
    )


def template_source_path(template_id: str, ext: str = ".pdf") -> Path:
    """Absolute path to a template's uploaded source (``.pdf`` by default, or ``.docx``)."""
    return _template_dir(template_id) / f"source{ext}"


def save_template_source(template_id: str, ext: str, content: bytes) -> Path:
    """Persist the raw source upload bytes as data/templates/<id>/source<ext>."""
    tmpl_dir = _template_dir(template_id)
    tmpl_dir.mkdir(parents=True, exist_ok=True)
    source = tmpl_dir / f"source{ext}"
    source.write_bytes(content)
    return source


def template_source_url(template_id: str, ext: str = ".pdf") -> str:
    """Relative URL (served via /files) for a template's source upload."""
    return f"/files/templates/{template_id}/source{ext}"


def template_outputs_dir(template_id: str) -> Path:
    """Directory holding a template's generated PDF outputs."""
    return _template_dir(template_id) / "outputs"


def save_template_output(
    template_id: str, output_id: str, content: bytes, ext: str = ".pdf"
) -> Path:
    """Persist a generated output as data/templates/<id>/outputs/<output_id><ext>."""
    out_dir = template_outputs_dir(template_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{output_id}{ext}"
    path.write_bytes(content)
    return path


def template_output_url(template_id: str, output_id: str, ext: str = ".pdf") -> str:
    """Relative URL (served via /files) for a template's generated output."""
    return f"/files/templates/{template_id}/outputs/{output_id}{ext}"


# --- Phase 4 (Vision QA): rendered + reference page images per QA run ----------

# A QA run rasterizes the template's preview PDF ("rendered") and, when available,
# the source/reference ("reference") into per-page PNGs the vision judge compares.


def template_qa_dir(template_id: str, run_id: str) -> Path:
    """Directory holding one QA run's rendered + reference page images."""
    return _template_dir(template_id) / "qa" / run_id


def save_qa_page(
    template_id: str, run_id: str, kind: str, page_no: int, content: bytes
) -> Path:
    """Write a QA page PNG as qa/<run_id>/<kind>/page-NNN.png (kind: "rendered"|"reference")."""
    out_dir = template_qa_dir(template_id, run_id) / kind
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"page-{page_no:03d}.png"
    path.write_bytes(content)
    return path


def qa_page_url(template_id: str, run_id: str, kind: str, page_no: int) -> str:
    """Relative URL (served via /files) for a QA run's rendered/reference page PNG."""
    return f"/files/templates/{template_id}/qa/{run_id}/{kind}/page-{page_no:03d}.png"


# --- Phase 6: outbound signing artifacts + demo signer certs -----------------


def original_path(doc_id: str) -> Path | None:
    """The saved original upload (data/<id>/original.*), or None if absent."""
    return next(iter(sorted(_doc_dir(doc_id).glob("original.*"))), None)


def read_original(doc_id: str) -> bytes:
    """Read the raw original upload bytes; raise if the file is missing."""
    path = original_path(doc_id)
    if path is None:
        raise FileNotFoundError(f"No original file on disk for document {doc_id}.")
    return path.read_bytes()


def signed_dir(doc_id: str) -> Path:
    """Directory holding the digitally signed output PDF(s)."""
    return _doc_dir(doc_id) / "signed"


def save_signed_pdf(doc_id: str, content: bytes, name: str = "signed.pdf") -> Path:
    """Write the signed PDF as signed/<name>."""
    out_dir = signed_dir(doc_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / name
    path.write_bytes(content)
    return path


def signed_pdf_path(doc_id: str, name: str = "signed.pdf") -> Path:
    """Absolute path to the signed PDF (may not exist yet)."""
    return signed_dir(doc_id) / name


def signed_pdf_exists(doc_id: str, name: str = "signed.pdf") -> bool:
    """True if a signed PDF has been written for this document."""
    return signed_pdf_path(doc_id, name).is_file()


def signed_pdf_url(doc_id: str, name: str = "signed.pdf") -> str:
    """Relative URL (served via /files) for the signed PDF."""
    return f"/files/{doc_id}/signed/{name}"


def delete_signed_dir(doc_id: str) -> None:
    """Remove data/<doc_id>/signed/ (the signed output PDF).

    Used to invalidate a stale signature when the decision it attested is re-run — a
    real signature must not outlive the approval it was based on. Idempotent: a missing
    directory is a no-op.
    """
    out_dir = signed_dir(doc_id)
    if out_dir.exists():
        shutil.rmtree(out_dir, ignore_errors=True)


def certs_dir() -> Path:
    """The demo signer cert dir (settings.signing_cert_path), created if absent.

    Kept OUTSIDE data_path so the /files mount never exposes the private keys.
    """
    path = settings.signing_cert_path
    path.mkdir(parents=True, exist_ok=True)
    return path
