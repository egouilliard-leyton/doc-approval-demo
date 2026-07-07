"""Spreadsheet preview: LibreOffice-headless formula recompute + computed grid / PDF.

openpyxl reads formula *strings*, never their results, and a cell we just wrote has no
cached value — so ``data_only=True`` reads it back as ``None``. LibreOffice headless is
the source of truth: convert the filled workbook (with a recalc-on-load profile) and
reopen it to read the computed values, or convert straight to PDF.

Everything here is a subprocess around the system ``soffice`` binary (no pip dep) and every
public function is a NON-RAISING, degrading boundary — the demo must never hard-fail on a
missing binary / non-zero exit / timeout / absent output:

- :func:`recompute_workbook` recomputes a filled workbook, disk-caching the result by
  ``sha256(xlsx_bytes)``; on ANY LibreOffice failure it returns the uncomputed input bytes
  flagged ``computed=False`` (never cached).
- :func:`convert_to_pdf` returns PDF bytes, or ``b""`` on any failure.
- :func:`read_computed_grid` builds the per-sheet grid from recomputed bytes; a formula cell
  whose value is still absent is flagged ``computed=False`` and shows its raw formula string.

LibreOffice isn't concurrency-safe on a shared profile, so each invocation runs with its own
``-env:UserInstallation`` (a throwaway seeded profile) and the calls are bounded by a
semaphore; the profile/work dir is always cleaned up.
"""

from __future__ import annotations

import hashlib
import logging
import shutil
import subprocess
import tempfile
import threading
from dataclasses import dataclass, field as dc_field
from io import BytesIO
from pathlib import Path

from app import storage
from app.config import settings
from app.schemas import SpreadsheetCell, SpreadsheetPreviewSheet

logger = logging.getLogger(__name__)

# Mirror the ingest-side caps so a pathological workbook can't produce a giant grid.
MAX_SHEET_ROWS = 500
MAX_SHEET_COLS = 60

# LibreOffice is not concurrency-safe even with isolated profiles; serialize (or bound)
# the soffice calls. Sized from settings at import so tests can tune concurrency via env.
_RECALC_SEMAPHORE = threading.Semaphore(max(1, int(settings.xlsx_recalc_concurrency)))

# A recalc-on-load profile: both ODF and OOXML recalc modes set to 0 == "Always
# recalculate on load". Without this, headless convert does NOT recompute and freshly
# written cells read back blank/zero. (Verified working in this environment.)
_REGISTRYMODIFICATIONS_XCU = """\
<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" \
xmlns:xs="http://www.w3.org/2001/XMLSchema" \
xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
</oor:items>
"""


@dataclass
class RecomputeResult:
    """Result of :func:`recompute_workbook`: the (possibly recomputed) bytes + status."""

    xlsx_bytes: bytes
    computed: bool  # True -> formulas recomputed by LibreOffice; False -> degraded input
    warnings: list[str] = dc_field(default_factory=list)
    cache_hit: bool = False


def _seed_profile(profile_dir: Path) -> None:
    """Write ``user/registrymodifications.xcu`` (recalc-on-load) into a LO profile dir."""
    user_dir = profile_dir / "user"
    user_dir.mkdir(parents=True, exist_ok=True)
    (user_dir / "registrymodifications.xcu").write_text(
        _REGISTRYMODIFICATIONS_XCU, encoding="utf-8"
    )


def _run_soffice(xlsx_bytes: bytes, convert_to: str, out_ext: str) -> tuple[bytes | None, str | None]:
    """Convert ``xlsx_bytes`` via headless soffice. Returns (output_bytes, warning).

    Runs with an isolated, recalc-seeded profile under a throwaway work dir (always cleaned
    up), bounded by the module semaphore and ``settings.xlsx_recalc_timeout_s``. Any failure
    — missing binary, non-zero exit, timeout, absent output — returns ``(None, reason)``
    rather than raising, so callers can degrade gracefully.
    """
    workdir = Path(tempfile.mkdtemp(prefix="xlsx_lo_"))
    profile_dir = workdir / "profile"
    outdir = workdir / "out"
    try:
        outdir.mkdir(parents=True, exist_ok=True)
        _seed_profile(profile_dir)
        input_path = workdir / "input.xlsx"
        input_path.write_bytes(xlsx_bytes)
        argv = [
            settings.xlsx_soffice_path,
            "--headless",
            "--calc",
            "--convert-to",
            convert_to,
            "--outdir",
            str(outdir),
            str(input_path),
            f"-env:UserInstallation=file://{profile_dir}",
        ]
        with _RECALC_SEMAPHORE:
            proc = subprocess.run(
                argv,
                capture_output=True,
                timeout=settings.xlsx_recalc_timeout_s,
            )
        out_path = outdir / f"input{out_ext}"
        if proc.returncode != 0:
            return None, f"soffice exited {proc.returncode}: {proc.stderr.decode(errors='replace')[:200]}"
        if not out_path.exists():
            return None, "soffice produced no output file"
        return out_path.read_bytes(), None
    except FileNotFoundError as exc:
        return None, f"soffice binary not found: {exc}"
    except subprocess.TimeoutExpired:
        return None, f"soffice timed out after {settings.xlsx_recalc_timeout_s}s"
    except Exception as exc:  # noqa: BLE001 — recompute is best-effort; never raise to callers
        return None, f"soffice recompute failed: {exc}"
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def recompute_workbook(template_id: str, xlsx_bytes: bytes) -> RecomputeResult:
    """Recompute a filled workbook's formulas via LibreOffice, disk-cached by content hash.

    Returns the recomputed bytes (``computed=True``) on success, caching them under
    ``data/templates/<id>/preview_cache/<sha256>.xlsx`` so paginating a preview never re-runs
    the (slow) recalc for identical bytes. On ANY LibreOffice failure it returns the
    *uncomputed input* bytes flagged ``computed=False`` with a warning — never raising, never
    caching the failure.
    """
    key = hashlib.sha256(xlsx_bytes).hexdigest()
    try:
        cache_path = storage.template_preview_cache_path(template_id, key)
        if cache_path.exists():
            return RecomputeResult(
                xlsx_bytes=cache_path.read_bytes(), computed=True, cache_hit=True
            )
    except Exception as exc:  # noqa: BLE001 — a cache hiccup just means "recompute afresh"
        logger.warning("preview cache lookup failed for %s: %s", template_id, exc)
        cache_path = None

    out_bytes, warning = _run_soffice(xlsx_bytes, "xlsx:Calc MS Excel 2007 XML", ".xlsx")
    if out_bytes is None:
        return RecomputeResult(
            xlsx_bytes=xlsx_bytes,
            computed=False,
            warnings=[f"formula recompute unavailable; showing raw formulas ({warning})"],
        )
    try:
        if cache_path is not None:
            cache_path.write_bytes(out_bytes)
    except Exception as exc:  # noqa: BLE001 — failing to cache doesn't invalidate the result
        logger.warning("preview cache write failed for %s: %s", template_id, exc)
    return RecomputeResult(xlsx_bytes=out_bytes, computed=True)


def convert_to_pdf(xlsx_bytes: bytes) -> bytes:
    """Convert a filled workbook to PDF via LibreOffice. Returns ``b""`` on any failure."""
    out_bytes, warning = _run_soffice(xlsx_bytes, "pdf:calc_pdf_Export", ".pdf")
    if out_bytes is None:
        logger.warning("xlsx->pdf conversion failed: %s", warning)
        return b""
    return out_bytes


def read_computed_grid(xlsx_bytes: bytes) -> list[SpreadsheetPreviewSheet]:
    """Build the per-sheet preview grid from (ideally recomputed) workbook bytes.

    Opens the workbook twice — ``data_only=True`` for computed values and ``data_only=False``
    for formula strings — so a formula cell that has a value renders it (``computed=True``)
    while one still lacking a value (LibreOffice recompute was unavailable) falls back to its
    raw formula string flagged ``computed=False``. A sheet is ``computed`` only if every
    formula cell on it resolved. Never raises: a total open failure returns ``[]``.
    """
    from openpyxl import load_workbook  # lazy

    try:
        wb_values = load_workbook(BytesIO(xlsx_bytes), data_only=True)
        wb_formulas = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    except Exception as exc:  # noqa: BLE001 — a corrupt workbook degrades to an empty preview
        logger.warning("preview grid read failed: %s", exc)
        return []

    try:
        sheets: list[SpreadsheetPreviewSheet] = []
        for name in wb_formulas.sheetnames:
            ws_f = wb_formulas[name]
            ws_v = wb_values[name] if name in wb_values.sheetnames else None
            row_cap = min(int(ws_f.max_row or 0), MAX_SHEET_ROWS)
            col_cap = min(int(ws_f.max_column or 0), MAX_SHEET_COLS)
            cells: list[SpreadsheetCell] = []
            sheet_computed = True
            for r in range(1, row_cap + 1):
                for c in range(1, col_cap + 1):
                    fcell = ws_f.cell(row=r, column=c)
                    raw = fcell.value
                    if raw is None:
                        continue
                    is_formula = isinstance(raw, str) and raw.startswith("=")
                    if is_formula:
                        value = ws_v.cell(row=r, column=c).value if ws_v is not None else None
                        if value is None:
                            computed = False
                            sheet_computed = False
                            display = raw  # raw formula string as the fallback display
                        else:
                            computed = True
                            display = _display(value)
                    else:
                        computed = True
                        display = _display(raw)
                    cells.append(
                        SpreadsheetCell(
                            row=r,
                            col=c,
                            address=fcell.coordinate,
                            value=display,
                            is_formula=is_formula,
                            number_format=fcell.number_format,
                            computed=computed,
                        )
                    )
            sheets.append(
                SpreadsheetPreviewSheet(
                    name=name,
                    max_row=row_cap,
                    max_col=col_cap,
                    merges=[str(rng) for rng in ws_f.merged_cells.ranges],
                    cells=cells,
                    computed=sheet_computed,
                )
            )
        return sheets
    except Exception as exc:  # noqa: BLE001 — never let preview rendering raise
        logger.warning("preview grid build failed: %s", exc)
        return []
    finally:
        wb_values.close()
        wb_formulas.close()


def _display(value: object) -> str | None:
    """Render a computed cell value as a display string (whole floats collapse to ints)."""
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)
