"""Spreadsheet templates (xlsx mode): enumerate, read, and fill a workbook.

The openpyxl core of the spreadsheet generation path. A spreadsheet template is a plain
``.xlsx`` whose formatting, formulas, charts, and column widths were authored beforehand;
the visual mapping UI binds catalogue fields to cells. This module:

- :func:`enumerate_workbook_sheets` — the per-sheet metadata captured at source-upload time
  (name / extent / merged ranges / column widths).
- :func:`read_template_grid` — a (capped) sheet grid for the mapping UI, showing formula
  strings verbatim (openpyxl reads formulas, not their results — recompute is LibreOffice's
  job, see :mod:`app.pipeline.generation.xlsx_preview`).
- :func:`fill_spreadsheet` — write a document's structured fields into the mapped cells:
  scalars into single cells (numeric suffix -> ``number_format``, text suffix -> literal
  concat), and list fields expanded down rows from an anchor (two row modes). The heavy
  formula recompute for preview/PDF lives in :mod:`xlsx_preview`; this module only writes.

Round-trips preserve values/formulas/number-formats/fonts/fills/borders/merges/col-widths
but LOSE charts, images, pivots, and cached formula results — the upload-time warning.
Pure openpyxl and offline; nothing here shells out.
"""

from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass, field as dc_field
from io import BytesIO
from pathlib import Path

from app import storage
from app.pipeline.generation.values import flatten_field_values, resolve_path
from app.schemas import SpreadsheetCell, SpreadsheetGrid, SpreadsheetSheetMeta

# Mirror the ingest-side caps (see ``storage._parse_xlsx``) so a pathological workbook
# can't produce a giant grid payload for the mapping UI.
MAX_SHEET_ROWS = 500
MAX_SHEET_COLS = 60


@dataclass
class SpreadsheetFillOutcome:
    """Result of :func:`fill_spreadsheet`: the filled bytes + a fill/skip trace."""

    content: bytes = b""
    filled: list[str] = dc_field(default_factory=list)  # written cell addresses (Sheet!A1)
    skipped: list[str] = dc_field(default_factory=list)  # bindings that wrote nothing
    warnings: list[str] = dc_field(default_factory=list)


def _cell_display(value: object) -> str | None:
    """Render a parsed cell as a display string; whole floats collapse to ints.

    Formula strings (``"=SUM(...)"``) are returned verbatim — the mapping grid shows the
    formula, not a (nonexistent) cached result. ``None`` stays ``None`` (an empty cell).
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def enumerate_workbook_sheets(path: str | Path) -> list[SpreadsheetSheetMeta]:
    """Per-sheet name / extent / merged ranges / column widths of a template workbook.

    Loaded non-``read_only`` (templates are small) so merged ranges and column widths are
    available. Persisted on the template at source-upload time so the mapping UI has the
    sheet layout without re-parsing the source each request.
    """
    from openpyxl import load_workbook  # lazy: keep import cost off app boot

    workbook = load_workbook(path, data_only=False)
    try:
        sheets: list[SpreadsheetSheetMeta] = []
        for ws in workbook.worksheets:
            col_widths = {
                letter: float(dim.width)
                for letter, dim in ws.column_dimensions.items()
                if dim.width is not None
            }
            sheets.append(
                SpreadsheetSheetMeta(
                    name=ws.title,
                    max_row=int(ws.max_row or 0),
                    max_col=int(ws.max_column or 0),
                    merges=[str(rng) for rng in ws.merged_cells.ranges],
                    col_widths=col_widths,
                )
            )
        return sheets
    finally:
        workbook.close()


def read_template_grid(
    path: str | Path,
    sheet: str,
    max_rows: int = MAX_SHEET_ROWS,
    max_cols: int = MAX_SHEET_COLS,
) -> SpreadsheetGrid:
    """A (capped) grid of one sheet's non-empty cells, for the click-to-bind mapping UI.

    Formula cells surface their formula string (``is_formula=True``); everything reads as a
    display string. Bounded to ``max_rows``/``max_cols`` like ``storage._parse_xlsx`` so the
    payload stays small. A missing sheet name yields an empty grid rather than raising.
    """
    from openpyxl import load_workbook  # lazy

    workbook = load_workbook(path, data_only=False)
    try:
        if sheet not in workbook.sheetnames:
            return SpreadsheetGrid(sheet=sheet, max_row=0, max_col=0, merges=[], cells=[])
        ws = workbook[sheet]
        row_cap = min(int(ws.max_row or 0), max_rows)
        col_cap = min(int(ws.max_column or 0), max_cols)
        cells: list[SpreadsheetCell] = []
        for r in range(1, row_cap + 1):
            for c in range(1, col_cap + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
                cells.append(
                    SpreadsheetCell(
                        row=r,
                        col=c,
                        address=cell.coordinate,
                        value=cell.value if is_formula else _cell_display(cell.value),
                        is_formula=is_formula,
                        number_format=cell.number_format,
                    )
                )
        return SpreadsheetGrid(
            sheet=sheet,
            max_row=row_cap,
            max_col=col_cap,
            merges=[str(rng) for rng in ws.merged_cells.ranges],
            cells=cells,
        )
    finally:
        workbook.close()


# --- fill --------------------------------------------------------------------


def _is_field_value(node: object) -> bool:
    """A dumped :class:`~app.schemas.FieldValue` leaf (value + confidence keys)."""
    return isinstance(node, dict) and "value" in node and "confidence" in node


def _record_value(record: object, field_path: str) -> object | None:
    """Resolve a table column's value out of one list record.

    ``field_path`` is record-relative: a sub-model field name for a ``list_composite`` row,
    or the ``""`` sentinel for a ``list_scalar`` row (the record *is* the value). Missing
    paths yield ``None`` rather than raising, mirroring :func:`resolve_path`.
    """
    if not field_path:
        return record["value"] if _is_field_value(record) else record
    return resolve_path(record, field_path)


def _apply_scalar(cell, value: object, suffix: str | None) -> None:
    """Write ``value`` into ``cell``, applying ``suffix`` per the value's kind.

    A numeric value keeps its number and gets a ``number_format`` unit (so downstream
    formulas still see a number); a text value with a suffix is literally concatenated.
    """
    is_numeric = isinstance(value, (int, float)) and not isinstance(value, bool)
    if suffix and is_numeric:
        cell.value = value
        cell.number_format = f'#,##0.00" {suffix}"'
    elif suffix:
        cell.value = f"{value} {suffix}"
    else:
        cell.value = value


def _clone_anchor_row(ws, anchor_row: int, dst_row: int) -> None:
    """Clone the anchor row's per-cell style + (translated) formula into ``dst_row``.

    openpyxl's ``insert_rows`` copies nothing — no style, no formula fix-up — so a freshly
    inserted row is blank/unstyled. This copies each anchor cell's ``_style`` and, for a
    formula cell, rewrites the formula relative to the new row via ``Translator`` (a
    non-formula value is copied verbatim). The mapped columns are overwritten afterward;
    unmapped columns (e.g. a per-row ``=qty*unit_price``) keep their cloned+translated form.
    """
    from openpyxl.formula.translate import Translator

    for cell in ws[anchor_row]:
        dst = ws.cell(row=dst_row, column=cell.column)
        dst._style = copy(cell._style)
        value = cell.value
        if isinstance(value, str) and value.startswith("="):
            try:
                dst.value = Translator(value, origin=cell.coordinate).translate_formula(
                    dst.coordinate
                )
            except Exception:  # noqa: BLE001 — a malformed formula copies verbatim, never crashes
                dst.value = value
        else:
            dst.value = value


# A same-sheet A1 cell-RANGE token (``X{r1}:Y{r2}``), with optional ``$`` absolute markers.
# The negative lookbehind keeps us off sheet-qualified refs (``Sheet1!A1:A5``) and off the
# tail of a longer identifier, so only bare same-sheet ranges are considered for expansion.
_RANGE_TOKEN_RE = re.compile(
    r"(?<![A-Za-z0-9_!$'.])"
    r"(\$?)([A-Za-z]{1,3})(\$?)(\d+)"
    r":"
    r"(\$?)([A-Za-z]{1,3})(\$?)(\d+)"
)


def _range_columns(col1: str, col2: str) -> tuple[int, int] | None:
    """The (lo, hi) 1-based column-index span of a range's two column letters, or None."""
    from openpyxl.utils import column_index_from_string

    try:
        i1 = column_index_from_string(col1.upper())
        i2 = column_index_from_string(col2.upper())
    except ValueError:
        return None
    return (min(i1, i2), max(i1, i2))


def _expand_table_formulas(
    ws, table_col_letters: set[str], anchor_row: int, inserted_count: int
) -> set[str]:
    """Widen bounded same-sheet ranges that ended at the anchor row to cover inserted rows.

    ``openpyxl.insert_rows`` shifts the POSITION of cells below the insertion point down but
    never rewrites their formula TEXT, so a bounded total under the table like ``=SUM(D4:D4)``
    keeps referencing only the single original anchor row and silently omits every inserted
    row. For each formula cell, rewrite every ``X{r1}:Y{r2}`` range token whose columns
    intersect the table columns AND whose END row equals ``anchor_row`` (i.e. it stopped at
    the pre-expansion anchor) so its end row becomes ``anchor_row + inserted_count``. Ranges
    are targeted by their ORIGINAL ``anchor_row`` reference — which does not move, the anchor
    being above the insertion point. Single-cell refs, other-sheet refs, and ranges not
    ending at the anchor are left untouched. Returns the set of rewritten cell coordinates.
    """
    expanded: set[str] = set()
    if inserted_count < 1 or not table_col_letters:
        return expanded
    from openpyxl.utils import column_index_from_string

    table_idx = {column_index_from_string(c.upper()) for c in table_col_letters}
    new_end = anchor_row + inserted_count

    def _rewrite(m: re.Match) -> str:
        c1a, c1, r1a, r1, c2a, c2, r2a, r2 = m.groups()
        span = _range_columns(c1, c2)
        if span is None:
            return m.group(0)
        lo, hi = span
        if not any(lo <= i <= hi for i in table_idx):
            return m.group(0)
        if int(r2) != anchor_row:
            return m.group(0)
        return f"{c1a}{c1}{r1a}{r1}:{c2a}{c2}{r2a}{new_end}"

    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not (isinstance(value, str) and value.startswith("=")):
                continue
            new_value = _RANGE_TOKEN_RE.sub(_rewrite, value)
            if new_value != value:
                cell.value = new_value
                expanded.add(cell.coordinate)
    return expanded


def _residual_formula_warnings(
    ws,
    table_col_letters: set[str],
    anchor_row: int,
    inserted_count: int,
    expanded: set[str],
    sheet: str,
    anchor: str,
) -> list[str]:
    """Best-effort warnings for below-table formulas step 1 could not auto-expand.

    A formula cell strictly below the (expanded) table region that references the table
    columns with a bounded range whose end row stops inside the pre-insert band
    (``anchor_row <= end < anchor_row + inserted_count``) may silently omit inserted rows.
    Whole-column ranges (``D:D``) carry no row digits and never match, so they never warn.
    """
    warnings: list[str] = []
    if inserted_count < 1 or not table_col_letters:
        return warnings
    from openpyxl.utils import column_index_from_string

    table_idx = {column_index_from_string(c.upper()) for c in table_col_letters}
    band_end = anchor_row + inserted_count
    for row in ws.iter_rows(min_row=band_end + 1):
        for cell in row:
            value = cell.value
            if not (isinstance(value, str) and value.startswith("=")):
                continue
            if cell.coordinate in expanded:
                continue
            risky = False
            for m in _RANGE_TOKEN_RE.finditer(value):
                span = _range_columns(m.group(2), m.group(6))
                if span is None:
                    continue
                lo, hi = span
                if not any(lo <= i <= hi for i in table_idx):
                    continue
                if anchor_row <= int(m.group(8)) < band_end:
                    risky = True
                    break
            if risky:
                warnings.append(
                    f"Inserted {inserted_count} row(s) at {sheet}!{anchor}; formula in "
                    f"{cell.coordinate} may not cover the new rows — verify in preview."
                )
    return warnings


def _fill_table(ws, table: dict, structured_fields: dict, max_rows: int) -> tuple[list[str], list[str]]:
    """Expand one table binding down rows from its anchor. Returns (filled, warnings)."""
    from openpyxl.utils import column_index_from_string, coordinate_to_tuple

    filled: list[str] = []
    warnings: list[str] = []

    list_path = table.get("list_path")
    anchor_cell = table.get("anchor_cell")
    if not list_path or not anchor_cell:
        return filled, warnings

    records = resolve_path(structured_fields, list_path)
    if not isinstance(records, list):
        warnings.append(f"{list_path}: not a list value; wrote 0 rows")
        return filled, warnings

    n = min(len(records), max_rows)
    if len(records) > max_rows:
        warnings.append(
            f"{list_path}: {len(records)} rows capped at {max_rows} (xlsx_max_table_rows)"
        )
    if n == 0:
        return filled, warnings

    anchor_row, _anchor_col = coordinate_to_tuple(anchor_cell)
    row_mode = table.get("row_mode") or "fill_next_empty_row"
    columns = table.get("columns", []) or []

    # insert_row: clone the styled anchor row (with translated formulas) into a fresh row
    # per record beyond the first, so per-row formulas + styling survive the expansion.
    inserted_count = 0
    if row_mode == "insert_row" and n > 1:
        inserted_count = n - 1
        ws.insert_rows(anchor_row + 1, amount=inserted_count)
        for i in range(1, n):
            _clone_anchor_row(ws, anchor_row, anchor_row + i)

    for i in range(n):
        record = records[i]
        target_row = anchor_row + i
        for col in columns:
            letter = col.get("col")
            if not letter:
                continue
            value = _record_value(record, col.get("field_path", ""))
            if value is None:
                continue
            cell = ws.cell(row=target_row, column=column_index_from_string(letter))
            _apply_scalar(cell, value, col.get("suffix"))
            filled.append(cell.coordinate)

    # After inserting rows + writing records, auto-expand bounded totals that ended at the
    # anchor row so they cover the inserted rows, then warn on any residual below-table
    # formula step 1 could not fix. (No-op unless the insert_row path actually inserted.)
    if inserted_count >= 1:
        table_cols = {col.get("col").upper() for col in columns if col.get("col")}
        for cell in ws[anchor_row]:
            if cell.value is not None:
                table_cols.add(cell.column_letter)
        expanded = _expand_table_formulas(ws, table_cols, anchor_row, inserted_count)
        warnings.extend(
            _residual_formula_warnings(
                ws, table_cols, anchor_row, inserted_count, expanded, ws.title, anchor_cell
            )
        )

    return filled, warnings


def fill_spreadsheet(template, structured_fields: dict) -> SpreadsheetFillOutcome:
    """Write a document's structured fields into ``template``'s mapped cells.

    ``template`` is the ORM row (its ``cell_map`` drives the binding; its ``source_ext``
    locates the uploaded ``.xlsx``); ``structured_fields`` is a dumped structuring ``fields``
    blob. Scalars write a single cell (numeric suffix -> ``number_format``, text suffix ->
    literal concat); tables expand a list field down rows from an anchor (``fill_next_empty_row``
    overwrites existing rows, ``insert_row`` inserts fresh styled rows). A None value, a
    missing sheet, or a non-list table source is skipped with a warning — never guessed.
    """
    from openpyxl import load_workbook  # lazy

    outcome = SpreadsheetFillOutcome()
    source_path = storage.template_source_path(template.id, template.source_ext or ".xlsx")
    workbook = load_workbook(source_path, data_only=False)
    try:
        cell_map = template.cell_map or {}
        flat = flatten_field_values(structured_fields)
        max_rows = _max_table_rows()

        for scalar in cell_map.get("scalars", []) or []:
            sheet = scalar.get("sheet")
            address = scalar.get("cell")
            field_path = scalar.get("field_path")
            if scalar.get("is_signature"):
                # Reserved: signature stamping onto cells is a later build; never a value.
                outcome.skipped.append(f"{sheet}!{address}")
                continue
            if not sheet or not address or not field_path:
                outcome.skipped.append(f"{sheet}!{address}")
                continue
            if sheet not in workbook.sheetnames:
                outcome.skipped.append(f"{sheet}!{address}")
                outcome.warnings.append(f"{sheet}!{address}: no such sheet")
                continue
            value = flat.get(field_path)
            if value is None:
                outcome.skipped.append(f"{sheet}!{address}")
                outcome.warnings.append(f"{sheet}!{address}: no value at '{field_path}'")
                continue
            _apply_scalar(workbook[sheet][address], value, scalar.get("suffix"))
            outcome.filled.append(f"{sheet}!{address}")

        for table in cell_map.get("tables", []) or []:
            sheet = table.get("sheet")
            if not sheet or sheet not in workbook.sheetnames:
                outcome.warnings.append(f"{sheet}: no such sheet for table binding")
                continue
            filled, warnings = _fill_table(workbook[sheet], table, structured_fields, max_rows)
            outcome.filled.extend(f"{sheet}!{addr}" for addr in filled)
            outcome.warnings.extend(warnings)

        buf = BytesIO()
        workbook.save(buf)
        outcome.content = buf.getvalue()
        return outcome
    finally:
        workbook.close()


def _max_table_rows() -> int:
    """The per-table row cap (``settings.xlsx_max_table_rows``), read lazily."""
    from app.config import settings

    return int(settings.xlsx_max_table_rows)
